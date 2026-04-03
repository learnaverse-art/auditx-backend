"""
AuditX — KPMG-Grade Excel Report Generator
Generates a fully formatted, professional forensic audit Excel report.
"""

import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime

# ── Colour palette (KPMG-inspired) ────────────────────────────────────────
C = {
    "navy":      "00338D",   # KPMG navy
    "blue":      "005EB8",   # KPMG blue
    "light_blue":"CCDFF5",
    "teal":      "00A3A1",
    "dark_gray": "333333",
    "mid_gray":  "666666",
    "light_gray":"F2F2F2",
    "white":     "FFFFFF",
    "red":       "C00000",
    "amber":     "FF8C00",
    "green":     "375623",
    "light_red": "FFE0E0",
    "light_amber":"FFF3CD",
    "light_green":"E2EFDA",
    "row_alt":   "EEF3FB",
    "border":    "BFBFBF",
    "section_bg":"E9EFF9",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="333333", italic=False, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    t = Side(style="thin", color=C["border"])
    n = None
    if sides == "all":
        return Border(left=t, right=t, top=t, bottom=t)
    if sides == "bottom":
        return Border(bottom=t)
    if sides == "top":
        return Border(top=t)
    if sides == "outer":
        return Border(left=t, right=t, top=t, bottom=t)
    return Border()

def thick_border_bottom():
    return Border(bottom=Side(style="medium", color=C["navy"]))

def money(v):
    if v is None or v == "": return "—"
    try:
        return f"₦{float(v):,.2f}"
    except:
        return str(v)

def pct(v, total):
    try:
        p = float(v) / float(total) * 100
        return f"{p:.1f}%"
    except:
        return "—"

# ── Sheet helpers ──────────────────────────────────────────────────────────
def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def cell(ws, row, col, value="", bold=False, size=11, color="333333",
         bg=None, h="left", v="center", wrap=False, italic=False,
         border=None, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if border:
        c.border = border
    if num_format:
        c.number_format = num_format
    return c

def header_row(ws, row, cols, values, bg=C["navy"], fg=C["white"],
               size=11, bold=True, height=22):
    ws.row_dimensions[row].height = height
    for i, (col, val) in enumerate(zip(cols, values)):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=bold, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = align(h="center" if i > 0 else "left", v="center")
        c.border = thin_border()

def section_title(ws, row, col, end_col, text, bg=C["navy"], fg=C["white"], height=24):
    ws.row_dimensions[row].height = height
    merge(ws, row, col, row, end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=13, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h="left", v="center")
    c.border = thin_border()

def sub_header(ws, row, col, end_col, text, bg=C["section_bg"], height=18):
    ws.row_dimensions[row].height = height
    merge(ws, row, col, row, end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=11, color=C["navy"])
    c.fill = fill(bg)
    c.alignment = align(h="left", v="center")
    c.border = thick_border_bottom()

def data_row(ws, row, cols, values, alt=False, height=16, borders=True, align_map=None):
    ws.row_dimensions[row].height = height
    bg = C["row_alt"] if alt else C["white"]
    for i, (col, val) in enumerate(zip(cols, values)):
        h = "left"
        if align_map and i in align_map:
            h = align_map[i]
        elif i > 0 and isinstance(val, (int, float)):
            h = "right"
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(size=10, color=C["dark_gray"])
        c.fill = fill(bg)
        c.alignment = align(h=h, v="center")
        if borders:
            c.border = thin_border()

def risk_badge_color(risk):
    r = str(risk).upper()
    if "HIGH" in r:
        return C["red"], C["light_red"]
    if "MEDIUM" in r or "MED" in r:
        return C["amber"], C["light_amber"]
    return C["green"], C["light_green"]

def blank_row(ws, row, col, end_col, height=8):
    ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════

def build_cover(wb, d):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"28","C":"32","D":"20","E":"18","F":"18"})

    # Top banner
    for r in range(1, 9):
        ws.row_dimensions[r].height = 18
    for r in [1,2,3,4,5,6,7,8]:
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = fill(C["navy"])

    merge(ws, 2, 2, 2, 5)
    c = ws.cell(row=2, column=2, value="FORENSIC AUDIT REPORT")
    c.font = font(bold=True, size=22, color=C["white"])
    c.alignment = align(h="left", v="center")

    merge(ws, 3, 2, 3, 5)
    c = ws.cell(row=3, column=2, value=d.get("companyName","") + ("  —  " + d.get("bankName","") if d.get("bankName") else ""))
    c.font = font(size=14, color=C["light_blue"])
    c.alignment = align(h="left", v="center")

    merge(ws, 4, 2, 4, 5)
    c = ws.cell(row=4, column=2, value=f"Prepared by AuditX Forensic Audit Agent  |  {datetime.now().strftime('%d %B %Y')}")
    c.font = font(size=10, color="AACCEE", italic=True)
    c.alignment = align(h="left", v="center")

    # Risk rating block
    risk = str(d.get("riskRating","MEDIUM")).upper().replace("🔴","").replace("🟡","").replace("🟢","").strip()
    rf, rb = risk_badge_color(risk)
    merge(ws, 6, 2, 7, 3)
    c = ws.cell(row=6, column=2, value=f"OVERALL RISK RATING:  {risk}")
    c.font = font(bold=True, size=14, color=C["white"])
    c.fill = fill(rf)
    c.alignment = align(h="center", v="center")
    c.border = thin_border()

    # Period + ID block
    merge(ws, 6, 4, 6, 5)
    ws.cell(row=6, column=4, value=f"Period: {d.get('period','—')}").font = font(size=10, color=C["light_blue"])
    ws.cell(row=6, column=4).alignment = align(h="left", v="center")
    merge(ws, 7, 4, 7, 5)
    ws.cell(row=7, column=4, value=f"Transactions: {d.get('totalTransactions','—')}").font = font(size=10, color=C["light_blue"])
    ws.cell(row=7, column=4).alignment = align(h="left", v="center")

    # ── Key Metrics table ──
    r = 10
    section_title(ws, r, 2, 5, "  KEY FINANCIAL METRICS", height=22); r+=1

    metrics = [
        ("Total Inflows (Deposits)",     d.get("totalInflows",0),    C["blue"],       C["white"]),
        ("Total Outflows (Withdrawals)",  d.get("totalOutflows",0),   "555555",        C["white"]),
        ("Net Cash Flow",                 (d.get("totalInflows",0) or 0) - (d.get("totalOutflows",0) or 0), C["teal"], C["white"]),
        ("Estimated Revenue (Adj.)",      d.get("estimatedRevenue",0),C["green"],      C["white"]),
        ("Total Identified Costs",        d.get("totalCosts",0),      C["red"],        C["white"]),
        ("Intercompany Transfers (deducted)", d.get("intercompanyTotal",0), "888888", C["white"]),
        ("Estimated Taxable Revenue",     d.get("taxableRevenue",0),  C["navy"],       C["white"]),
        ("VAT Obligation (7.5%)",         d.get("vatObligation",0),   C["navy"],       C["white"]),
        ("CIT Estimate (30%)",            d.get("citEstimate",0),     C["navy"],       C["white"]),
    ]
    for label, val, fg, bg in metrics:
        ws.row_dimensions[r].height = 20
        merge(ws, r, 2, r, 3)
        c2 = ws.cell(row=r, column=2, value=label)
        c2.font = font(size=11, color=C["dark_gray"]); c2.fill = fill(C["light_gray"]); c2.alignment = align(); c2.border = thin_border()
        merge(ws, r, 4, r, 5)
        c4 = ws.cell(row=r, column=4, value=money(val))
        c4.font = font(bold=True, size=11, color=fg); c4.fill = fill(C["light_gray"]); c4.alignment = align(h="right"); c4.border = thin_border()
        r += 1

    # ── Red flags count ──
    r += 1
    ws.row_dimensions[r].height = 20
    merge(ws, r, 2, r, 3)
    c2 = ws.cell(row=r, column=2, value="Red Flags Identified")
    c2.font = font(bold=True, size=11, color=C["white"]); c2.fill = fill(C["red"]); c2.alignment = align(); c2.border = thin_border()
    merge(ws, r, 4, r, 5)
    c4 = ws.cell(row=r, column=4, value=d.get("redFlagCount", len(d.get("redFlags",[])))); c4.font = font(bold=True, size=14, color=C["red"]); c4.fill = fill(C["light_red"]); c4.alignment = align(h="center"); c4.border = thin_border()
    r += 2

    # ── Contents table ──
    section_title(ws, r, 2, 5, "  REPORT CONTENTS", height=22); r+=1
    contents = [
        ("1", "Executive Summary",       "Key findings, overview and risk rating"),
        ("2", "Scope & Methodology",     "Audit approach, data sources and limitations"),
        ("3", "Account Profile",         "Account details and balance analysis"),
        ("4", "Transaction Ledger",      "Significant transactions with categorization and flags"),
        ("5", "Revenue & Cost Analysis", "Adjusted revenue, cost breakdown, profitability, tax"),
        ("6", "Exception Report",        "Red flags with severity, amounts and recommended actions"),
        ("7", "Risk Assessment",         "Fraud, AML, tax risks and action plan"),
        ("8", "Appendices",              "Supporting data and category summaries"),
    ]
    for num, title, desc in contents:
        ws.row_dimensions[r].height = 18
        cell(ws,r,2,num, bold=True, color=C["navy"], bg=C["section_bg"], border=thin_border())
        merge(ws, r, 3, r, 3)
        cell(ws,r,3,title, bold=True, color=C["dark_gray"], bg=C["white"], border=thin_border())
        merge(ws, r, 4, r, 5)
        cell(ws,r,4,desc, color=C["mid_gray"], bg=C["white"], border=thin_border())
        r += 1

    r += 1
    merge(ws, r, 2, r, 5)
    c = ws.cell(row=r, column=2, value="⚠️ Colour Legend:   Yellow = Estimated / Uncertain   |   Red = High Risk / Requires Immediate Action   |   Green = Verified / Low Risk")
    c.font = font(size=9, color=C["mid_gray"], italic=True)
    c.alignment = align(h="left", v="center", wrap=True)
    ws.row_dimensions[r].height = 20

    # Freeze & protect presentation
    ws.sheet_view.showRowColHeaders = False
    return ws


def build_exec_summary(wb, d):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"22","C":"18","D":"18","E":"18","F":"22"})

    r = 2
    section_title(ws, r, 2, 6, "  1. EXECUTIVE SUMMARY", height=26); r+=2

    # Overview paragraph
    sub_header(ws, r, 2, 6, "1.1  Overview"); r+=1
    ws.row_dimensions[r].height = 80
    merge(ws, r, 2, r, 6)
    c = ws.cell(row=r, column=2, value=d.get("overview","Forensic analysis completed. See individual sections for detailed findings."))
    c.font = font(size=10.5, color=C["dark_gray"])
    c.alignment = align(h="left", v="top", wrap=True)
    c.fill = fill(C["white"])
    c.border = thin_border()
    r += 2

    # Risk summary boxes
    sub_header(ws, r, 2, 6, "1.2  Risk Summary"); r+=1
    risk_items = [
        ("Overall Risk",       d.get("riskRating","—")),
        ("Fraud Risk",         d.get("fraudRisk","—")),
        ("AML Risk",           d.get("amlRisk","—")),
        ("Tax Compliance Risk",d.get("taxRisk","—")),
    ]
    cols_used = [2,3,4,5]
    for col, (label, val) in zip(cols_used, risk_items):
        ws.row_dimensions[r].height = 14
        cell(ws,r,col,label, bold=True, size=9, color=C["white"], bg=C["navy"], h="center", border=thin_border())
    r+=1
    for col, (label, val) in zip(cols_used, risk_items):
        raw = str(val).upper().replace("🔴","").replace("🟡","").replace("🟢","").strip()
        fc, bc = risk_badge_color(raw)
        ws.row_dimensions[r].height = 22
        cell(ws,r,col,raw, bold=True, size=12, color=fc, bg=bc, h="center", border=thin_border())
    r += 2

    # Key findings
    sub_header(ws, r, 2, 6, "1.3  Key Findings"); r+=1
    findings = d.get("keyFindings", d.get("findings", []))
    if not findings:
        findings = ["No key findings extracted — review individual sections."]
    for i, f in enumerate(findings):
        ws.row_dimensions[r].height = max(18, min(60, len(str(f))//4))
        merge(ws,r,2,r,2)
        cell(ws,r,2,f"  {i+1}.", bold=True, color=C["navy"], bg=C["row_alt"] if i%2 else C["white"], border=thin_border())
        merge(ws,r,3,r,6)
        c = ws.cell(row=r, column=3, value=str(f))
        c.font = font(size=10, color=C["dark_gray"])
        c.alignment = align(h="left", v="center", wrap=True)
        c.fill = fill(C["row_alt"] if i%2 else C["white"])
        c.border = thin_border()
        r += 1
    r += 1

    # Recommendations
    sub_header(ws, r, 2, 6, "1.4  Recommendations"); r+=1
    recs = d.get("recommendations", [])
    if not recs:
        recs = ["Obtain supporting documentation for all flagged transactions.",
                "Engage a qualified tax advisor for full tax compliance review."]
    for i, rec in enumerate(recs):
        ws.row_dimensions[r].height = max(18, min(60, len(str(rec))//4))
        merge(ws,r,2,r,2)
        cell(ws,r,2,f"  {i+1}.", bold=True, color=C["navy"], bg=C["section_bg"], border=thin_border())
        merge(ws,r,3,r,6)
        c = ws.cell(row=r, column=3, value=str(rec))
        c.font = font(size=10, color=C["dark_gray"])
        c.alignment = align(h="left", v="center", wrap=True)
        c.fill = fill(C["section_bg"])
        c.border = thin_border()
        r += 1

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_scope(wb, d):
    ws = wb.create_sheet("Scope & Methodology")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"28","C":"40","D":"20"})
    r = 2
    section_title(ws, r, 2, 4, "  2. SCOPE & METHODOLOGY", height=26); r+=2

    sub_header(ws, r, 2, 4, "2.1  Scope of Audit"); r+=1
    scope_rows = [
        ("Data Source",            d.get("bankName","Bank statement provided")),
        ("Company",                d.get("companyName","—")),
        ("Period Covered",         d.get("period","—")),
        ("Total Transactions",     str(d.get("totalTransactions","—"))),
        ("Opening Balance",        money(d.get("openingBalance",0))),
        ("Closing Balance",        money(d.get("closingBalance",0))),
        ("Materiality Threshold",  "₦500,000"),
        ("Currency",               "Nigerian Naira (NGN)"),
    ]
    for i,(label,val) in enumerate(scope_rows):
        alt = i%2==1
        ws.row_dimensions[r].height = 18
        cell(ws,r,2,label, bold=True, size=10, color=C["navy"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        merge(ws,r,3,r,4)
        cell(ws,r,3,val, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        r+=1
    r+=1

    sub_header(ws, r, 2, 4, "2.2  Methodology"); r+=1
    methods = [
        "Data Extraction: Bank statement data extracted and normalised for analysis.",
        "Transaction Categorisation: Automated classification based on descriptions and patterns.",
        "Revenue Identification: Exclusion of intercompany transfers, loan receipts, CBN forex conversions.",
        "Cost Analysis: Classification of outflows into operating cost categories.",
        "Anomaly Detection: Statistical analysis to identify outliers and unusual patterns.",
        "Red Flag Analysis: Identification of high-risk transactions requiring further investigation.",
    ]
    for i, m in enumerate(methods):
        ws.row_dimensions[r].height = 24
        merge(ws,r,2,r,4)
        c = ws.cell(row=r, column=2, value=f"  • {m}")
        c.font = font(size=10, color=C["dark_gray"])
        c.alignment = align(h="left", v="center", wrap=True)
        c.fill = fill(C["row_alt"] if i%2 else C["white"])
        c.border = thin_border()
        r+=1
    r+=1

    sub_header(ws, r, 2, 4, "2.3  Items Excluded from Revenue"); r+=1
    exclusions = [
        ("Intercompany Transfers",  money(d.get("intercompanyTotal",0)),  "Internal fund movements between related entities"),
        ("Loan Receipts",           money(d.get("loansReceived",0)),       "Debt instruments — not business revenue"),
        ("CBN/Forex Conversions",   money(d.get("forexConversions",0)),    "Currency conversion — not operating income"),
        ("Returned Payments",       "—",                                   "Reversals and refunded transactions"),
    ]
    header_row(ws, r, [2,3,4], ["Item","Amount Deducted","Rationale"]); r+=1
    for i,(item,amt,reason) in enumerate(exclusions):
        alt=i%2==1
        ws.row_dimensions[r].height = 18
        cell(ws,r,2,item, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        cell(ws,r,3,amt, size=10, color=C["red"], bg=C["row_alt"] if alt else C["white"], border=thin_border(), h="right")
        cell(ws,r,4,reason, size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        r+=1

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_account_profile(wb, d):
    ws = wb.create_sheet("Account Profile")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"28","C":"30","D":"22","E":"22"})
    r = 2
    section_title(ws, r, 2, 5, "  3. ACCOUNT PROFILE", height=26); r+=2

    sub_header(ws, r, 2, 5, "3.1  Account Information"); r+=1
    info = [
        ("Account Name",       d.get("companyName","—")),
        ("Bank",               d.get("bankName","—")),
        ("Account Type",       d.get("accountType","Current Account")),
        ("Currency",           "Nigerian Naira (NGN)"),
        ("Period Analysed",    d.get("period","—")),
    ]
    for i,(k,v) in enumerate(info):
        alt=i%2==1
        ws.row_dimensions[r].height = 20
        cell(ws,r,2,k, bold=True, size=10, color=C["navy"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        merge(ws,r,3,r,5)
        cell(ws,r,3,v, size=10.5, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        r+=1
    r+=1

    sub_header(ws, r, 2, 5, "3.2  Balance Summary"); r+=1
    ti = d.get("totalInflows",0) or 0
    to_ = d.get("totalOutflows",0) or 0
    balance_rows = [
        ("Opening Balance",    money(d.get("openingBalance",0)),       C["mid_gray"]),
        ("Total Deposits",     money(ti),                              C["green"]),
        ("Total Withdrawals",  money(to_),                             C["red"]),
        ("Net Cash Flow",      money(ti - to_),                       C["blue"]),
        ("Closing Balance",    money(d.get("closingBalance",0)),       C["navy"]),
    ]
    for i,(label,val,col) in enumerate(balance_rows):
        alt=i%2==1
        ws.row_dimensions[r].height = 22
        cell(ws,r,2,label, bold=True, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        merge(ws,r,3,r,5)
        cell(ws,r,3,val, bold=True, size=12, color=col, bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        r+=1
    r+=1

    sub_header(ws, r, 2, 5, "3.3  Transaction Volume"); r+=1
    ws.row_dimensions[r].height = 20
    cell(ws,r,2,"Total Transactions", bold=True, size=10, color=C["dark_gray"], bg=C["light_gray"], border=thin_border())
    merge(ws,r,3,r,5)
    cell(ws,r,3,str(d.get("totalTransactions","—")), bold=True, size=12, color=C["navy"], h="right", bg=C["light_gray"], border=thin_border())
    r+=2

    # Mini chart placeholder note
    merge(ws,r,2,r,5)
    c = ws.cell(row=r, column=2, value="Note: Detailed transaction categorisation and volume analysis available in the Transaction Ledger and Appendices sheets.")
    c.font = font(size=9, color=C["mid_gray"], italic=True)
    c.alignment = align(h="left", v="center", wrap=True)
    ws.row_dimensions[r].height = 24

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_transaction_ledger(wb, d):
    ws = wb.create_sheet("Transaction Ledger")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"13","C":"13","D":"16","E":"40","F":"16","G":"16","H":"18","I":"22","J":"14","K":"16"})
    r = 2
    section_title(ws, r, 2, 11, "  4. TRANSACTION LEDGER — Significant Transactions", height=26); r+=2

    summary_note = d.get("transactionLedgerSummary", "Chronological listing of significant transactions. Flags: 🔴 High Risk | 🟡 Verify | ✅ Clean | 🔵 Deducted (intercompany / loan)")
    merge(ws,r,2,r,11)
    c = ws.cell(row=r, column=2, value=summary_note)
    c.font = font(size=10, color=C["mid_gray"], italic=True)
    c.alignment = align(h="left", v="center", wrap=True)
    c.fill = fill(C["section_bg"])
    c.border = thin_border()
    ws.row_dimensions[r].height = 30
    r+=2

    # Headers
    header_row(ws, r, list(range(2,12)),
               ["Date","Value Date","Reference","Description","Withdrawal (₦)","Deposit (₦)","Balance (₦)","Category","Party Type","Flag"],
               bg=C["navy"], height=22); r+=1

    txns = d.get("significantTransactions", d.get("transactions", []))
    if not txns:
        merge(ws,r,2,r,11)
        cell(ws,r,2,"No transaction detail extracted. Paste raw bank statement into Agent Mode for line-by-line analysis.",
             italic=True, color=C["mid_gray"], bg=C["light_gray"], border=thin_border())
        ws.row_dimensions[r].height=20; r+=1
    else:
        for i,tx in enumerate(txns[:50]):
            alt=i%2==1
            bg=C["row_alt"] if alt else C["white"]
            flag=str(tx.get("flag","")).lower()
            if "🔴" in flag or "high" in flag:
                bg=C["light_red"]
            elif "🟡" in flag or "verify" in flag:
                bg=C["light_amber"]
            elif "🔵" in flag or "deduct" in flag or "interco" in flag:
                bg="EEF3FB"
            ws.row_dimensions[r].height=18
            vals = [
                tx.get("date",""),
                tx.get("valueDate",tx.get("date","")),
                tx.get("reference",""),
                tx.get("description",""),
                money(tx.get("withdrawal",tx.get("debit",None))) if tx.get("withdrawal") or tx.get("debit") else "",
                money(tx.get("deposit",tx.get("credit",None))) if tx.get("deposit") or tx.get("credit") else "",
                money(tx.get("balance",tx.get("runningBalance",None))),
                tx.get("category",""),
                tx.get("partyType",""),
                tx.get("flag",""),
            ]
            for j,(col,val) in enumerate(zip(range(2,12),vals)):
                c = ws.cell(row=r, column=col, value=val)
                c.font = font(size=10, color=C["dark_gray"])
                c.fill = fill(bg)
                c.alignment = align(h="right" if j in [4,5,6] else "left", v="center")
                c.border = thin_border()
            r+=1

    # Deductions summary
    r+=1
    sub_header(ws, r, 2, 11, "4.1  Items Deducted (not counted as revenue/costs)"); r+=1
    deducted = d.get("deductedItems", d.get("deductions", {}))
    if isinstance(deducted, dict):
        deducted_list = [(k.title(), v, "") for k,v in deducted.items() if v]
    else:
        deducted_list = deducted or []
    if deducted_list:
        header_row(ws, r, [2,3,4,5], ["Category","Total Amount Deducted","Count","Reason"],
                   bg=C["teal"]); r+=1
        for i,item in enumerate(deducted_list):
            alt=i%2==1
            if isinstance(item, dict):
                cat=item.get("category",""); amt=item.get("totalAmount",item.get("amount",0)); cnt=item.get("count",""); reason=item.get("reason","")
            else:
                cat,amt,reason=item[0],item[1],""; cnt=""
            ws.row_dimensions[r].height=18
            cell(ws,r,2,cat, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
            cell(ws,r,3,money(amt), bold=True, size=10, color=C["green"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
            cell(ws,r,4,str(cnt), size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], h="center", border=thin_border())
            cell(ws,r,5,reason, size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
            r+=1

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_revenue_cost(wb, d):
    ws = wb.create_sheet("Revenue & Cost Analysis")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"32","C":"20","D":"10","E":"14","F":"20"})
    r = 2
    section_title(ws, r, 2, 6, "  5. REVENUE & COST ANALYSIS", height=26); r+=2

    # Revenue
    sub_header(ws, r, 2, 6, "5.1  Revenue Estimation (After Deductions)"); r+=1
    header_row(ws, r, [2,3,4,5,6], ["Revenue Category","Amount (₦)","Count","% of Total","Notes"],
               bg=C["teal"]); r+=1
    rev_items = d.get("revenueItems", [])
    if not rev_items:
        rev_items = [{"category":"Estimated Revenue (Adjusted)", "amount": d.get("estimatedRevenue",0),
                      "count":"—","notes":"After deducting intercompany, loans, forex"}]
    total_rev = d.get("estimatedRevenue",0) or sum(x.get("amount",0) for x in rev_items)
    for i,item in enumerate(rev_items):
        alt=i%2==1
        amt = item.get("amount",0)
        ws.row_dimensions[r].height=18
        cell(ws,r,2,item.get("category",""), size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        cell(ws,r,3,money(amt), bold=True, size=10, color=C["green"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        cell(ws,r,4,str(item.get("count","—")), size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], h="center", border=thin_border())
        cell(ws,r,5,pct(amt,total_rev) if total_rev else "—", size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], h="center", border=thin_border())
        cell(ws,r,6,item.get("notes",""), size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border(), wrap=True)
        r+=1
    # Total row
    ws.row_dimensions[r].height=20
    cell(ws,r,2,"ADJUSTED REVENUE (Conservative)", bold=True, size=11, color=C["white"], bg=C["teal"], border=thin_border())
    merge(ws,r,3,r,3)
    cell(ws,r,3,money(d.get("estimatedRevenue",total_rev)), bold=True, size=11, color=C["white"], bg=C["teal"], h="right", border=thin_border())
    for col in [4,5,6]: cell(ws,r,col,"", bg=C["teal"], border=thin_border())
    r+=2

    # Cost Breakdown
    sub_header(ws, r, 2, 6, "5.2  Cost Analysis"); r+=1
    header_row(ws, r, [2,3,4,5,6], ["Cost Category","Amount (₦)","Count","% of Total","Notes"],
               bg=C["navy"]); r+=1
    costs = d.get("costBreakdown",[])
    if not costs:
        costs = [{"category":"Costs not extracted","amount":d.get("totalCosts",0),"count":"—","notes":"Re-run with detailed bank statement"}]
    total_cost = d.get("totalCosts",0) or sum(c.get("amount",0) for c in costs if isinstance(c,dict))
    for i,item in enumerate(costs):
        if not isinstance(item,dict): continue
        alt=i%2==1
        amt=item.get("amount",0)
        ws.row_dimensions[r].height=18
        cell(ws,r,2,item.get("category",""), size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        cell(ws,r,3,money(amt), bold=True, size=10, color=C["red"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        cell(ws,r,4,str(item.get("count","—")), size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], h="center", border=thin_border())
        cell(ws,r,5,pct(amt,total_cost) if total_cost else "—", size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], h="center", border=thin_border())
        cell(ws,r,6,item.get("notes",""), size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border(), wrap=True)
        r+=1
    ws.row_dimensions[r].height=20
    cell(ws,r,2,"TOTAL IDENTIFIED COSTS", bold=True, size=11, color=C["white"], bg=C["navy"], border=thin_border())
    cell(ws,r,3,money(total_cost), bold=True, size=11, color=C["white"], bg=C["navy"], h="right", border=thin_border())
    for col in [4,5,6]: cell(ws,r,col,"", bg=C["navy"], border=thin_border())
    r+=2

    # Profitability
    sub_header(ws, r, 2, 6, "5.3  Profitability Summary"); r+=1
    rev = d.get("estimatedRevenue",0) or 0
    cost= d.get("totalCosts",0) or 0
    net = rev - cost
    prof_rows = [
        ("Estimated Revenue",       money(rev),  C["green"]),
        ("Less: Total Costs",       f"({money(cost)})", C["red"]),
        ("Net Operating Position",  money(net),  C["blue"] if net>=0 else C["red"]),
    ]
    for i,(label,val,col) in enumerate(prof_rows):
        ws.row_dimensions[r].height=22
        cell(ws,r,2,label, bold=(i==2), size=11, color=C["dark_gray"], bg=C["section_bg"], border=thin_border())
        merge(ws,r,3,r,6)
        cell(ws,r,3,val, bold=(i==2), size=12, color=col, bg=C["section_bg"], h="right", border=thin_border())
        r+=1
    r+=2

    # Tax
    sub_header(ws, r, 2, 6, "5.4  Tax Implications (FIRS / NRS)"); r+=1
    header_row(ws, r, [2,3,4,5], ["Tax Head","Estimated Amount (₦)","Basis","Confidence"],
               bg=C["navy"]); r+=1
    tax_rows = [
        ("Taxable Revenue",   d.get("taxableRevenue",0),   "Adjusted revenue after deductions",          d.get("taxConfidence","⚠️ Estimated")),
        ("VAT (7.5%)",        d.get("vatObligation",0),    "7.5% of taxable revenue",                   "⚠️ Estimated"),
        ("CIT (30%)",         d.get("citEstimate",0),      "30% Corporate Income Tax",                  "⚠️ Estimated"),
        ("WHT",               d.get("wht",0),              "Withholding Tax on applicable transactions","🔴 Uncertain"),
    ]
    for i,(label,amt,basis,conf) in enumerate(tax_rows):
        alt=i%2==1
        conf_bg = C["light_amber"] if "Estim" in str(conf) else C["light_red"] if "Uncert" in str(conf) else C["light_green"]
        ws.row_dimensions[r].height=18
        cell(ws,r,2,label, bold=True, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        cell(ws,r,3,money(amt), bold=True, size=11, color=C["navy"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        cell(ws,r,4,basis, size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        cell(ws,r,5,conf, size=10, color=C["dark_gray"], bg=conf_bg, h="center", border=thin_border())
        r+=1
    r+=1
    merge(ws,r,2,r,5)
    c=ws.cell(row=r,column=2,value="⚠️ All tax figures are estimates based on available data and require verification by a qualified FIRS-registered tax advisor.")
    c.font=font(size=9,color=C["mid_gray"],italic=True)
    c.alignment=align(h="left",v="center",wrap=True)
    c.fill=fill(C["light_amber"])
    ws.row_dimensions[r].height=24

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_exception_report(wb, d):
    ws = wb.create_sheet("Exception Report")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"30","C":"12","D":"20","E":"12","F":"36","G":"30"})
    r = 2
    section_title(ws, r, 2, 7, "  6. EXCEPTION REPORT — RED FLAGS", height=26); r+=2

    flags = d.get("redFlags", d.get("exceptionReport", []))
    total_flags = len(flags) if flags else d.get("redFlagCount",0)
    merge(ws,r,2,r,7)
    cell(ws,r,2,f"Total Flags Identified: {total_flags}  |  Review each flag and obtain supporting documentation.",
         bold=True, color=C["red"], bg=C["light_red"], border=thin_border())
    ws.row_dimensions[r].height=20; r+=2

    header_row(ws, r, [2,3,4,5,6,7],
               ["Flag Description","Count","Amount (₦)","Risk Level","Detailed Description","Recommended Action"],
               bg=C["red"], fg=C["white"], height=24); r+=1

    if not flags:
        merge(ws,r,2,r,7)
        cell(ws,r,2,"No exception data extracted — re-run with detailed bank statement.",
             italic=True, color=C["mid_gray"], bg=C["light_gray"], border=thin_border())
        ws.row_dimensions[r].height=20; r+=1
    else:
        for i, fl in enumerate(flags):
            if not isinstance(fl, dict): continue
            risk=str(fl.get("risk","Medium"))
            fc, bc = risk_badge_color(risk)
            h_scale = max(30, min(80, len(str(fl.get("description","")))/2))
            ws.row_dimensions[r].height=h_scale

            def exc_cell(col, val, h="left"):
                c2 = ws.cell(row=r, column=col, value=val)
                c2.font = font(size=10, color=C["dark_gray"])
                c2.fill = fill(bc+"44" if bc else C["white"])
                c2.alignment = align(h=h, v="top", wrap=True)
                c2.border = thin_border()
            exc_cell(2, fl.get("flag", fl.get("description","—")))
            exc_cell(3, str(fl.get("count","—")), "center")
            exc_cell(4, money(fl.get("amount",0)), "right")
            rc = ws.cell(row=r, column=5, value=risk.upper())
            rc.font = font(bold=True, size=10, color=fc)
            rc.fill = fill(bc)
            rc.alignment = align(h="center", v="center")
            rc.border = thin_border()
            exc_cell(6, fl.get("description",""))
            exc_cell(7, fl.get("action",fl.get("recommendedAction","Obtain supporting documentation and verify.")))
            r+=1

    r+=2
    sub_header(ws, r, 2, 7, "6.1  AML Concerns"); r+=1
    aml = d.get("amlConcerns", [])
    if not aml:
        aml = ["No specific AML concerns extracted — review intercompany and round-sum transactions."]
    for i, concern in enumerate(aml):
        alt=i%2==1
        ws.row_dimensions[r].height=max(18,min(60,len(str(concern))//3))
        merge(ws,r,2,r,7)
        c=ws.cell(row=r,column=2,value=f"  • {concern}")
        c.font=font(size=10,color=C["dark_gray"])
        c.alignment=align(h="left",v="center",wrap=True)
        c.fill=fill(C["row_alt"] if alt else C["white"])
        c.border=thin_border()
        r+=1

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_risk_assessment(wb, d):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"26","C":"14","D":"36","E":"28"})
    r = 2
    section_title(ws, r, 2, 5, "  7. RISK ASSESSMENT", height=26); r+=2

    # Overall
    sub_header(ws, r, 2, 5, "7.1  Overall Risk Rating"); r+=1
    risk=str(d.get("riskRating","—")).upper().replace("🔴","").replace("🟡","").replace("🟢","").strip()
    fc, bc = risk_badge_color(risk)
    ws.row_dimensions[r].height=32
    merge(ws,r,2,r,5)
    c=ws.cell(row=r,column=2,value=f"  {risk}")
    c.font=font(bold=True,size=16,color=fc)
    c.fill=fill(bc)
    c.alignment=align(h="left",v="center")
    c.border=thin_border()
    r+=2

    sub_header(ws, r, 2, 5, "7.2  Risk Categories"); r+=1
    header_row(ws, r, [2,3,4,5], ["Risk Area","Level","Description","Mitigation"], bg=C["navy"]); r+=1
    risk_rows = [
        ("Fraud Risk",              d.get("fraudRisk","—"),   d.get("fraudDesc","Review all flagged transactions for authenticity."),   "Independent verification of major transactions."),
        ("AML Risk",                d.get("amlRisk","—"),     d.get("amlDesc","High intercompany activity requires scrutiny."),          "Obtain transfer pricing documentation."),
        ("Revenue Recognition",     d.get("revenueRisk","—"), d.get("revenueRiskDesc","Revenue may be understated or overstated."),      "Review all bank accounts and revenue policies."),
        ("Tax Compliance",          d.get("taxRisk","—"),     d.get("taxRiskDesc","Related party transactions may need disclosure."),    "Engage qualified FIRS-registered tax advisor."),
        ("Control Environment",     "Medium",                  "Duplicate references and unusual patterns may indicate control gaps.",    "Strengthen transaction approval and audit trail."),
    ]
    for i,(label,level,desc,mit) in enumerate(risk_rows):
        lv=str(level).upper().replace("🔴","").replace("🟡","").replace("🟢","").strip()
        fc2, bc2 = risk_badge_color(lv)
        alt=i%2==1
        h_row=max(28,min(60,len(str(desc))//3))
        ws.row_dimensions[r].height=h_row
        cell(ws,r,2,label, bold=True, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        lc=ws.cell(row=r,column=3,value=lv if lv!="—" else "—")
        lc.font=font(bold=True,size=10,color=fc2); lc.fill=fill(bc2); lc.alignment=align(h="center",v="center"); lc.border=thin_border()
        cell(ws,r,4,desc, size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border(), wrap=True)
        cell(ws,r,5,mit, size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border(), wrap=True)
        r+=1
    r+=2

    # Action Plan
    sub_header(ws, r, 2, 5, "7.3  Recommended Action Plan"); r+=1
    action_sections=[
        ("IMMEDIATE ACTIONS (Now — within 7 days)", C["red"],   d.get("immediateActions",["Obtain transfer pricing docs for intercompany > ₦10M","Verify all related party transactions with supporting docs","Reconcile import costs with sales revenue"])),
        ("SHORT-TERM ACTIONS (30 days)",             C["amber"], d.get("shortTermActions",["Investigate duplicate reference numbers","Review mobile banking transaction records","Obtain Form M and PAAR documentation"])),
        ("LONG-TERM ACTIONS (90 days)",              C["teal"],  d.get("longTermActions",["Implement unique reference numbering system","Strengthen transaction approval process","Commission full transfer pricing study"])),
    ]
    for timeframe, col, actions in action_sections:
        ws.row_dimensions[r].height=20
        merge(ws,r,2,r,5)
        c=ws.cell(row=r,column=2,value=f"  {timeframe}")
        c.font=font(bold=True,size=11,color=C["white"])
        c.fill=fill(col); c.alignment=align(h="left",v="center"); c.border=thin_border()
        r+=1
        for i,action in enumerate(actions or []):
            alt=i%2==1
            ws.row_dimensions[r].height=max(18,min(50,len(str(action))//3))
            merge(ws,r,2,r,5)
            c=ws.cell(row=r,column=2,value=f"  • {action}")
            c.font=font(size=10,color=C["dark_gray"])
            c.alignment=align(h="left",v="center",wrap=True)
            c.fill=fill(C["row_alt"] if alt else C["white"])
            c.border=thin_border()
            r+=1
        r+=1

    ws.sheet_view.showRowColHeaders = False
    return ws


def build_appendices(wb, d):
    ws = wb.create_sheet("Appendices")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":"4","B":"28","C":"18","D":"18","E":"16","F":"12"})
    r = 2
    section_title(ws, r, 2, 6, "  8. APPENDICES", height=26); r+=2

    sub_header(ws, r, 2, 6, "Appendix A — Transaction Category Summary"); r+=1
    header_row(ws, r, [2,3,4,5,6], ["Category","Withdrawals (₦)","Deposits (₦)","Net (₦)","Count"],
               bg=C["navy"]); r+=1
    cats = d.get("categoryAppendix", d.get("costBreakdown", []))
    if not cats:
        cats = [{"category": "Data not available", "withdrawals":0,"deposits":0,"net":0,"count":"—"}]
    for i,cat in enumerate(cats):
        if not isinstance(cat,dict): continue
        alt=i%2==1
        ws.row_dimensions[r].height=18
        cell(ws,r,2,cat.get("category",""), size=10, color=C["dark_gray"], bg=C["row_alt"] if alt else C["white"], border=thin_border())
        cell(ws,r,3,money(cat.get("withdrawals",cat.get("amount",0))), size=10, color=C["red"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        cell(ws,r,4,money(cat.get("deposits",0)), size=10, color=C["green"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        cell(ws,r,5,money(cat.get("net",0)), size=10, color=C["blue"], bg=C["row_alt"] if alt else C["white"], h="right", border=thin_border())
        cell(ws,r,6,str(cat.get("count","—")), size=10, color=C["mid_gray"], bg=C["row_alt"] if alt else C["white"], h="center", border=thin_border())
        r+=1
    r+=2

    sub_header(ws, r, 2, 6, "Appendix B — Confidence Legend"); r+=1
    legend=[
        ("✅ High Confidence",  C["green"],       C["light_green"],  "Figures verified from available data with high certainty"),
        ("⚠️ Estimated",        C["amber"],       C["light_amber"],  "Figures estimated — require verification with supporting documents"),
        ("🔴 Uncertain",        C["red"],         C["light_red"],    "Figures uncertain — treat as indicative only"),
    ]
    for sym, fc2, bc2, desc in legend:
        ws.row_dimensions[r].height=20
        cell(ws,r,2,sym, bold=True, size=10, color=fc2, bg=bc2, border=thin_border())
        merge(ws,r,3,r,6)
        cell(ws,r,3,desc, size=10, color=C["dark_gray"], bg=bc2, border=thin_border())
        r+=1
    r+=2

    merge(ws,r,2,r,6)
    c=ws.cell(row=r,column=2,value="This report was generated by AuditX Forensic Audit Agent. All figures should be independently verified. This report does not constitute a formal audit opinion and is intended for preliminary analysis only.")
    c.font=font(size=9,color=C["mid_gray"],italic=True)
    c.alignment=align(h="left",v="center",wrap=True)
    c.fill=fill(C["light_gray"])
    c.border=thin_border()
    ws.row_dimensions[r].height=36

    ws.sheet_view.showRowColHeaders = False
    return ws


# ══════════════════════════════════════════════════════════════════════════
# MAIN GENERATOR
# ══════════════════════════════════════════════════════════════════════════

def generate_report(data: dict, output_path: str = None) -> str:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_cover(wb, data)
    build_exec_summary(wb, data)
    build_scope(wb, data)
    build_account_profile(wb, data)
    build_transaction_ledger(wb, data)
    build_revenue_cost(wb, data)
    build_exception_report(wb, data)
    build_risk_assessment(wb, data)
    build_appendices(wb, data)

    company = data.get("companyName","Audit").replace(" ","_")[:20]
    if not output_path:
        output_path = f"/mnt/user-data/outputs/{company}_KPMG_Report.xlsx"
    wb.save(output_path)
    return output_path


# ── Test with RELCHEM data ─────────────────────────────────────────────────
if __name__ == "__main__":
    # Load sample data from command line or use test data
    if len(sys.argv) > 1:
        with open(sys.argv[1]) as f:
            data = json.load(f)
    else:
        data = {
            "companyName": "RELCHEM LIMITED",
            "bankName": "First Bank of Nigeria",
            "period": "03-Jan-2024 to 04-Apr-2025",
            "riskRating": "HIGH",
            "totalInflows": 4584280843.80,
            "totalOutflows": 4552710011.38,
            "estimatedRevenue": 724085000.00,
            "totalCosts": 2345435093.12,
            "intercompanyTotal": 4291720000.00,
            "loansReceived": 0,
            "forexConversions": 0,
            "vatObligation": 54306375.00,
            "citEstimate": 217225500.00,
            "wht": 0,
            "taxableRevenue": 724085000.00,
            "openingBalance": 99950.00,
            "closingBalance": 31670782.42,
            "totalTransactions": 1710,
            "accountType": "Current Account (Corporate)",
            "taxConfidence": "⚠️ Estimated",
            "overview": "This forensic audit report presents findings from a comprehensive analysis of RELCHEM LIMITED's bank statement for the period 03-January-2024 to 04-April-2025. Total inflows of ₦4.58 billion were recorded against outflows of ₦4.55 billion, yielding a net cash flow of ₦31.57 million. After deducting intercompany transfers (₦4.29 billion), estimated revenue stands at ₦724 million against identified costs of ₦2.35 billion, indicating that import costs significantly exceed identified revenue — a critical finding requiring explanation. Risk is rated HIGH due to high intercompany activity, duplicate reference numbers, and related party transaction volumes.",
            "riskFactors": [
                "Import costs (₦2.22B) significantly exceed identified revenue (₦724M)",
                "High volume of intercompany transfers (₦4.29B) across 194 transactions",
                "Multiple related party transactions (₦1.09B) requiring disclosure",
                "Large round-number transactions (₦6.04B) indicating possible structuring",
                "Duplicate reference numbers (385 instances) suggesting system issues",
            ],
            "recommendations": [
                "Obtain transfer pricing documentation for all intercompany transactions exceeding ₦10M",
                "Verify business purpose and obtain supporting documents for all related party transactions",
                "Reconcile import costs (₦2.22B) with sales revenue and inventory records",
                "Investigate duplicate reference numbers and obtain explanations for each instance",
                "Review mobile banking transaction records for proper audit trail",
            ],
            "fraudRisk": "HIGH",
            "amlRisk": "HIGH",
            "taxRisk": "MEDIUM",
            "revenueRisk": "HIGH",
            "fraudDesc": "Large round-number transactions, duplicate references, and unexplained related party transactions suggest potential manipulation.",
            "amlDesc": "High volume of intercompany transfers and related party transactions without clear business purpose.",
            "revenueRiskDesc": "Import costs significantly exceed identified revenue — revenue may be recognised through other channels or bank accounts.",
            "taxRiskDesc": "Related party transactions may not be properly disclosed for transfer pricing and CIT purposes.",
            "costBreakdown": [
                {"category":"Customs Duty Payments",       "amount":1168693403.29,"count":54,"notes":"49.8% of total costs"},
                {"category":"BCF Import Settlement",       "amount":1051558750.00,"count":2, "notes":"44.8% of total costs"},
                {"category":"Insurance",                   "amount":82341227.01,  "count":4, "notes":"3.5% of total costs"},
                {"category":"Personnel (Payroll)",         "amount":22782398.98,  "count":3, "notes":"Salary payments"},
                {"category":"Shipping & Logistics",        "amount":14154996.88,  "count":8, "notes":"Freight and clearing"},
                {"category":"Taxes (LIRS)",                "amount":3552216.96,   "count":2, "notes":"Lagos IRS payments"},
                {"category":"Operating Expenses",          "amount":1941500.00,   "count":5, "notes":"Misc. operating costs"},
                {"category":"Import Processing (Form M)",  "amount":410600.00,    "count":2, "notes":"Import documentation"},
            ],
            "redFlags": [
                {"flag":"Large Round-Number Transactions", "count":214,"amount":6042163000.00,"risk":"Medium","description":"Transactions with round figures (e.g., ₦10M, ₦50M) may indicate structuring or lack of proper documentation.","action":"Obtain supporting invoices and business rationale for all round-sum transactions > ₦5M."},
                {"flag":"High Intercompany Activity",      "count":194,"amount":4291720000.00,"risk":"High",  "description":"Significant intercompany transfers require transfer pricing documentation and arm's length verification.","action":"Obtain transfer pricing study and legal agreements for all intercompany transactions."},
                {"flag":"Related Party Transactions",      "count":212,"amount":1089034000.00,"risk":"High",  "description":"Multiple related party transactions requiring disclosure and substantiation.","action":"Verify business purpose; obtain board resolutions and arm's-length confirmation."},
                {"flag":"Duplicate Reference Numbers",     "count":385,"amount":569606399.96, "risk":"Medium","description":"Duplicate references may indicate split transactions, system errors, or potential manipulation.","action":"Request bank confirmation and investigate each duplicate systematically."},
                {"flag":"High Mobile Banking Volume",      "count":102,"amount":428780000.00, "risk":"Medium","description":"Significant mobile banking deposits may indicate cash collection without proper audit trail.","action":"Obtain mobile banking records and trace to customer invoices."},
                {"flag":"Import Costs Exceed Revenue",     "count":54, "amount":2220662753.29,"risk":"High",  "description":"Import-related costs (₦2.22B) significantly exceed customer revenue (₦724M).","action":"Reconcile with inventory records, review other bank accounts, and obtain sales ledger."},
            ],
            "amlConcerns": [
                "High volume of intercompany transfers (₦4.29B across 194 transactions) without clear documented business purpose",
                "Multiple related party transactions through various individuals (Ododo Blessing, Agudile Foster, Okoro Emeka) requiring KYC review",
                "Round-figure transactions totalling ₦6.04B may indicate layering or structuring",
                "Mobile banking inflows of ₦428M without traceable invoices raise cash placement concerns",
            ],
            "immediateActions": [
                "Obtain transfer pricing documentation for all intercompany transactions > ₦10M",
                "Verify business purpose and obtain supporting documents for all related party transactions",
                "Reconcile import costs (₦2.22B) with sales revenue, purchase orders and inventory",
            ],
            "shortTermActions": [
                "Investigate all 385 duplicate reference numbers and obtain bank explanations",
                "Review mobile banking transaction records for proper audit trail",
                "Obtain Form M and PAAR documentation to verify import legitimacy",
            ],
            "longTermActions": [
                "Commission a full transfer pricing study covering all intercompany transactions",
                "Implement unique reference numbering and strengthen transaction approval processes",
                "Engage FIRS-registered tax advisor for comprehensive tax compliance review",
            ],
            "significantTransactions": [
                {"date":"02-Sep-2024","description":"FIP:ZIB/RELCHEM LIMITED","type":"Credit","deposit":10000000,"balance":None,"category":"Intercompany","flag":"🔵 Deducted"},
                {"date":"02-Sep-2024","description":"FIP:UBA/RELCHEM LIMITED/CIB/UTO","type":"Credit","deposit":38000000,"balance":None,"category":"Intercompany","flag":"🔵 Deducted"},
                {"date":"05-Sep-2024","description":"FIP:ZIB/RELCHEM LIMITED/FIRST BANK","type":"Credit","deposit":40000000,"balance":None,"category":"Intercompany","flag":"🔵 Deducted"},
            ],
        }

    out = generate_report(data)
    print(f"✅ Report generated: {out}")
